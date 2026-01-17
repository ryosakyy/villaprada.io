# backend/routers/reportes.py

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from datetime import date
from typing import List

from fastapi.responses import StreamingResponse

from core.database import get_db
from core.security import get_current_user
from models.pagos import Pago
from models.egresos import Egreso
from models.contratos import Contrato
from models.clientes import Cliente

import io
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

router = APIRouter(
    prefix="/reportes",
    tags=["Reportes"],
)


# ============================================================
# 🧾 API REAL QUE USA TU FRONTEND → /reportes/finanzas
# ============================================================
@router.get("/finanzas", dependencies=[Depends(get_current_user)])
def reporte_finanzas(inicio: date, fin: date, db: Session = Depends(get_db)):

    if fin < inicio:
        raise HTTPException(status_code=400, detail="El rango de fechas es inválido")

    datos = []

    fecha_actual = inicio
    while fecha_actual <= fin:

        ingresos = (
            db.query(Pago.monto)
            .filter(Pago.fecha_pago == fecha_actual)
            .all()
        )
        total_ingresos = sum(m[0] for m in ingresos) if ingresos else 0

        egresos = (
            db.query(Egreso.monto)
            .filter(Egreso.fecha == fecha_actual)
            .all()
        )
        total_egresos = sum(m[0] for m in egresos) if egresos else 0

        saldo = total_ingresos - total_egresos

        datos.append({
            "fecha": str(fecha_actual),
            "ingresos": total_ingresos,
            "egresos": total_egresos,
            "saldo": saldo
        })

        fecha_actual = date.fromordinal(fecha_actual.toordinal() + 1)

    return datos


# ============================================================
# 1) REPORTE EXCEL DE INGRESOS
# ============================================================
@router.get("/ingresos/excel", dependencies=[Depends(get_current_user)])
def reporte_ingresos_excel(fecha_inicio: date, fecha_fin: date, db: Session = Depends(get_db)):

    if fecha_fin < fecha_inicio:
        raise HTTPException(status_code=400, detail="Rango inválido")

    pagos = (
        db.query(Pago)
        .filter(Pago.fecha_pago >= fecha_inicio, Pago.fecha_pago <= fecha_fin)
        .order_by(Pago.fecha_pago.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ingresos"

    ws.append(["ID", "Fecha pago", "Contrato ID", "Monto", "Método", "Observación"])

    total = 0
    for p in pagos:
        ws.append([
            p.id,
            str(p.fecha_pago),
            p.contrato_id,
            float(p.monto),
            p.metodo,
            p.observacion or ""
        ])
        total += float(p.monto)

    ws.append([])
    ws.append(["", "", "TOTAL", total])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ingresos_{fecha_inicio}_{fecha_fin}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ============================================================
# 2) REPORTE EXCEL DE EGRESOS
# ============================================================
@router.get("/egresos/excel", dependencies=[Depends(get_current_user)])
def reporte_egresos_excel(fecha_inicio: date, fecha_fin: date, db: Session = Depends(get_db)):

    if fecha_fin < fecha_inicio:
        raise HTTPException(status_code=400, detail="Rango inválido")

    egresos = (
        db.query(Egreso)
        .filter(Egreso.fecha >= fecha_inicio, Egreso.fecha <= fecha_fin)
        .order_by(Egreso.fecha.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Egresos"

    ws.append(["ID", "Fecha", "Descripción", "Categoría", "Monto", "Contrato ID"])

    total = 0
    for e in egresos:
        ws.append([
            e.id,
            e.fecha,
            e.descripcion,
            e.categoria,
            float(e.monto),
            e.contrato_id or ""
        ])
        total += float(e.monto)

    ws.append([])
    ws.append(["", "", "", "TOTAL", total])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"egresos_{fecha_inicio}_{fecha_fin}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ============================================================
# 3) REPORTE EXCEL DE CONTRATOS
# ============================================================
@router.get("/contratos/excel", dependencies=[Depends(get_current_user)])
def reporte_contratos_excel(fecha_inicio: date, fecha_fin: date, db: Session = Depends(get_db)):

    if fecha_fin < fecha_inicio:
        raise HTTPException(status_code=400, detail="Rango inválido")

    contratos = (
        db.query(Contrato)
        .filter(Contrato.fecha_evento >= fecha_inicio, Contrato.fecha_evento <= fecha_fin)
        .order_by(Contrato.fecha_evento.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"

    ws.append([
        "ID", "Cliente", "Fecha evento", "Paquete", "Monto total",
        "Adelanto", "Saldo", "Estado"
    ])

    for c in contratos:
        cliente = c.cliente.nombres if c.cliente else "SIN CLIENTE"

        ws.append([
            c.id,
            cliente,
            c.fecha_evento,
            c.paquete,
            float(c.monto_total),
            float(c.adelanto),
            float(c.saldo),
            c.estado
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"contratos_{fecha_inicio}_{fecha_fin}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ============================================================
# 4) REPORTE PDF — FLUJO DE CAJA
# ============================================================
@router.get("/flujo-caja/pdf", dependencies=[Depends(get_current_user)])
def reporte_flujo_caja_pdf(fecha_inicio: date, fecha_fin: date, db: Session = Depends(get_db)):

    if fecha_fin < fecha_inicio:
        raise HTTPException(status_code=400, detail="Rango inválido")

    pagos = db.query(Pago).filter(Pago.fecha_pago >= fecha_inicio, Pago.fecha_pago <= fecha_fin).all()
    egresos = db.query(Egreso).filter(Egreso.fecha >= fecha_inicio, Egreso.fecha <= fecha_fin).all()

    total_ingresos = sum(float(p.monto) for p in pagos)
    total_egresos = sum(float(e.monto) for e in egresos)
    utilidad = total_ingresos - total_egresos

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 50

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y, "Reporte de Flujo de Caja - Villa Prado")
    y -= 30

    c.setFont("Helvetica", 11)
    c.drawString(50, y, f"Periodo: {fecha_inicio} al {fecha_fin}")
    y -= 20

    c.drawString(50, y, f"Ingresos: S/ {total_ingresos:.2f}")
    y -= 15
    c.drawString(50, y, f"Egresos: S/ {total_egresos:.2f}")
    y -= 15
    c.drawString(50, y, f"UTILIDAD: S/ {utilidad:.2f}")
    y -= 30

    # Detalle de ingresos / egresos
    # ...

    c.save()
    buffer.seek(0)

    filename = f"flujocaja_{fecha_inicio}_{fecha_fin}.pdf"

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
